import string
import random
import uuid
import re
from io import BytesIO
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr, field_validator
from sqlalchemy.orm import Session
from sqlalchemy import or_, desc, asc
import openpyxl

from app.core.dependencies import get_db
from app.core.rbac_middleware import PermissionChecker
from app.core.auth_middleware import get_current_user_no_password_force
from app.models.models import User, Role, Permission, UserPermission, Department, Section, Designation, UserProfile, AuditLog, UserStatus
from app.core.security import get_password_hash
from app.core.responses import make_response

router = APIRouter()

# ----------------------------------------------------
# Request / Response Schemas
# ----------------------------------------------------
class UserCreateRequest(BaseModel):
    email: EmailStr
    name: str
    role_name: str
    department_id: str | None = None
    section_id: str | None = None
    permissions: list[str] | None = None

class UserUpdateRequest(BaseModel):
    name: str
    email: EmailStr
    department_id: str | None = None
    section_id: str | None = None
    phone_number: str | None = None
    bio: str | None = None
    address: str | None = None
    designation_id: str | None = None

    @field_validator("email")
    @classmethod
    def validate_email_domain(cls, v: str) -> str:
        # Require standard email address structures
        regex = r"^[\w\.-]+@[\w\.-]+\.\w+$"
        if not re.match(regex, v):
            raise ValueError("Invalid email format.")
        return v

class StatusUpdateRequest(BaseModel):
    user_ids: list[str]
    status: UserStatus

# Helper generators
def generate_username(name: str) -> str:
    cleaned = "".join(c for c in name.lower() if c.isalnum())
    if not cleaned:
        cleaned = "user"
    rand_suffix = "".join(random.choices(string.digits, k=4))
    return f"{cleaned}{rand_suffix}"

def generate_temp_password() -> str:
    # Ensure temporary password meets complexity checks
    lower = random.choice(string.ascii_lowercase)
    upper = random.choice(string.ascii_uppercase)
    digit = random.choice(string.digits)
    spec = random.choice("!@#$%^&*")
    others = "".join(random.choices(string.ascii_letters + string.digits + "!@#$%^&*", k=8))
    return f"{lower}{upper}{digit}{spec}{others}"

# ----------------------------------------------------
# Endpoints
# ----------------------------------------------------

@router.get("", dependencies=[Depends(PermissionChecker("users:read"))])
def list_users(
    page: int = 1,
    limit: int = 10,
    search: str | None = None,
    role: str | None = None,
    status: str | None = None,
    department_id: str | None = None,
    sort_by: str = "createdAt",
    sort_order: str = "desc",
    db: Session = Depends(get_db)
):
    """Lists users with pagination, filters, searching, and sorting options."""
    query = db.query(User)

    # 1. Searching
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            or_(
                User.name.like(search_filter),
                User.email.like(search_filter),
                User.username.like(search_filter)
            )
        )

    # 2. Filtering
    if role:
        query = query.join(User.role).filter(Role.name == role)
    if status:
        query = query.filter(User.status == status)
    if department_id:
        query = query.filter(User.departmentId == department_id)

    # 3. Sorting
    sort_column = getattr(User, sort_by, User.createdAt)
    if sort_order.lower() == "desc":
        query = query.order_by(desc(sort_column))
    else:
        query = query.order_by(asc(sort_column))

    # 4. Pagination calculations
    total = query.count()
    offset = (page - 1) * limit
    users = query.offset(offset).limit(limit).all()

    users_list = [
        {
            "id": u.id,
            "email": u.email,
            "username": u.username,
            "name": u.name,
            "role": u.role.name,
            "status": u.status,
            "department": u.department.name if u.department else None,
            "section": u.section.name if u.section else None,
            "createdAt": u.createdAt
        } for u in users
    ]
    res_data = {
        "users": users_list,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }
    return make_response(
        success=True,
        message="Users listed successfully.",
        data=res_data,
        extra_compat=res_data
    )

@router.post("", dependencies=[Depends(PermissionChecker("users:create"))])
def create_user(payload: UserCreateRequest, current_user: User = Depends(get_current_user_no_password_force), db: Session = Depends(get_db)):
    """Provisions a new staff or student user account."""
    if payload.role_name == "MASTER_ADMIN":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot provision duplicate MASTER_ADMIN accounts."
        )

    role = db.query(Role).filter_by(name=payload.role_name).first()
    if not role:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Role '{payload.role_name}' not resolved.")

    if db.query(User).filter_by(email=payload.email).first():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email already linked to another account.")

    # Resolve mappings
    if payload.department_id:
        dept = db.query(Department).filter_by(id=payload.department_id).first()
        if not dept:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Department not resolved.")

    if payload.section_id:
        sect = db.query(Section).filter_by(id=payload.section_id).first()
        if not sect:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Section not resolved.")

    username = generate_username(payload.name)
    while db.query(User).filter_by(username=username).first():
        username = generate_username(payload.name)

    temp_pwd = generate_temp_password()
    hashed = get_password_hash(temp_pwd)

    user_id = str(uuid.uuid4())
    user = User(
        id=user_id,
        email=payload.email,
        username=username,
        passwordHash=hashed,
        name=payload.name,
        roleId=role.id,
        departmentId=payload.department_id,
        sectionId=payload.section_id,
        status="ACTIVE",
        mustChangePassword=True
    )
    db.add(user)

    # Initialize associated profile
    profile = UserProfile(id=str(uuid.uuid4()), userId=user_id)
    db.add(profile)

    # Bind direct permissions
    if payload.permissions:
        for perm_name in payload.permissions:
            perm = db.query(Permission).filter_by(name=perm_name).first()
            if perm:
                link = UserPermission(id=str(uuid.uuid4()), userId=user_id, permissionId=perm.id)
                db.add(link)

    # Audit log
    audit = AuditLog(
        id=str(uuid.uuid4()),
        userId=current_user.id,
        action="USER_CREATE",
        details=f"Created user {username} ({payload.email}) with role {payload.role_name}."
    )
    db.add(audit)
    db.commit()

    print(f"\n[MOCK EMAIL SERVICE] To: {payload.email}\nUsername: {username}\nTemp Password: {temp_pwd}\n")

    user_data = {
        "user": {
            "id": user_id,
            "email": user.email,
            "username": username,
            "name": user.name,
            "role": payload.role_name
        },
        "temporary_password": temp_pwd
    }
    return make_response(
        success=True,
        message="User created successfully.",
        data=user_data,
        extra_compat=user_data
    )

@router.get("/{id}", dependencies=[Depends(PermissionChecker("users:read"))])
def get_user(id: str, db: Session = Depends(get_db)):
    """Retrieves full profile data, designation, and direct permissions of a user."""
    user = db.query(User).filter_by(id=id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User account not resolved.")

    # Load custom permission list
    perms = [up.permission.name for up in user.userPermissions]

    user_data = {
        "user": {
            "id": user.id,
            "email": user.email,
            "username": user.username,
            "name": user.name,
            "role": user.role.name,
            "status": user.status,
            "departmentId": user.departmentId,
            "department": user.department.name if user.department else None,
            "sectionId": user.sectionId,
            "section": user.section.name if user.section else None,
            "mustChangePassword": user.mustChangePassword,
            "createdAt": user.createdAt,
            "profile": {
                "phoneNumber": user.profile.phoneNumber if user.profile else None,
                "bio": user.profile.bio if user.profile else None,
                "address": user.profile.address if user.profile else None,
                "avatarUrl": user.profile.avatarUrl if user.profile else None,
                "designationId": user.profile.designationId if (user.profile and user.profile.designationId) else None,
                "designation": user.profile.designation.name if (user.profile and user.profile.designation) else None
            },
            "permissions": perms
        }
    }
    return make_response(
        success=True,
        message="User retrieved successfully.",
        data=user_data,
        extra_compat=user_data
    )

@router.put("/{id}", dependencies=[Depends(PermissionChecker("users:update"))])
def update_user(id: str, payload: UserUpdateRequest, current_user: User = Depends(get_current_user_no_password_force), db: Session = Depends(get_db)):
    """Updates user information and extends profile classifications."""
    user = db.query(User).filter_by(id=id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User account not resolved.")

    # Email uniqueness verification
    if payload.email != user.email:
        duplicate = db.query(User).filter_by(email=payload.email).first()
        if duplicate:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email already mapped to another user.")

    # Master Admin protection
    if user.role.name == "MASTER_ADMIN" and payload.email != user.email:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Modifying email profiles on the MASTER_ADMIN is prohibited."
        )

    # Map core properties
    user.name = payload.name
    user.email = payload.email
    user.departmentId = payload.department_id
    user.sectionId = payload.section_id

    # Verify relationships
    if payload.section_id:
        sect = db.query(Section).filter_by(id=payload.section_id).first()
        if not sect:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Section not found.")

    if not user.profile:
        user.profile = UserProfile(id=str(uuid.uuid4()), userId=user.id)

    user.profile.phoneNumber = payload.phone_number
    user.profile.bio = payload.bio
    user.profile.address = payload.address
    user.profile.designationId = payload.designation_id

    if payload.designation_id:
        desg = db.query(Designation).filter_by(id=payload.designation_id).first()
        if not desg:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Designation not resolved.")

    # Audit log
    audit = AuditLog(
        id=str(uuid.uuid4()),
        userId=current_user.id,
        action="USER_UPDATE",
        details=f"Modified profile parameters for user {user.username}."
    )
    db.add(audit)
    db.commit()

    return make_response(
        success=True,
        message="User updated successfully.",
        data={}
    )

@router.delete("/{id}", dependencies=[Depends(PermissionChecker("users:delete"))])
def delete_user(id: str, current_user: User = Depends(get_current_user_no_password_force), db: Session = Depends(get_db)):
    """Deletes a user account from database records."""
    user = db.query(User).filter_by(id=id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User account not resolved.")

    # Master Admin Protection
    if user.role.name == "MASTER_ADMIN":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Security Lock: The MASTER_ADMIN account cannot be deleted."
        )

    db.delete(user)

    audit = AuditLog(
        id=str(uuid.uuid4()),
        userId=current_user.id,
        action="USER_DELETE",
        details=f"Deleted user account: {user.username} ({user.email})."
    )
    db.add(audit)
    db.commit()

    return make_response(
        success=True,
        message="User account successfully purged.",
        data={}
    )

@router.patch("/status", dependencies=[Depends(PermissionChecker("users:update"))])
def update_status(payload: StatusUpdateRequest, current_user: User = Depends(get_current_user_no_password_force), db: Session = Depends(get_db)):
    """Adjusts account statuses for multiple target user profiles."""
    updated_count = 0
    for uid in payload.user_ids:
        user = db.query(User).filter_by(id=uid).first()
        if not user:
            continue
            
        # Protect Master Admin from updates
        if user.role.name == "MASTER_ADMIN":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Security Lock: MASTER_ADMIN status cannot be updated."
            )

        user.status = payload.status.value
        user.isSuspended = (payload.status == UserStatus.SUSPENDED)
        user.isDisabled = (payload.status == UserStatus.DISABLED)
        updated_count += 1

    audit = AuditLog(
        id=str(uuid.uuid4()),
        userId=current_user.id,
        action="USERS_STATUS_UPDATE",
        details=f"Set status to {payload.status.value} for {updated_count} user(s)."
    )
    db.add(audit)
    db.commit()

    return make_response(
        success=True,
        message=f"Updated status of {updated_count} accounts successfully.",
        data={}
    )

@router.post("/import", dependencies=[Depends(PermissionChecker("users:create"))])
async def import_users(file: UploadFile = File(...), current_user: User = Depends(get_current_user_no_password_force), db: Session = Depends(get_db)):
    """Parses Excel data spreadsheets and bulk provisions campus accounts."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported file format. Please upload an Excel sheet.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Failed to load spreadsheet: {str(e)}")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Spreadsheet contains no data rows.")

    headers = [str(h).strip().lower() for h in rows[0]]
    
    # Required columns map
    required_cols = ["name", "email", "role"]
    for col in required_cols:
        if col not in headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Required header '{col}' missing. Must contain 'Name', 'Email', and 'Role'."
            )

    idx_name = headers.index("name")
    idx_email = headers.index("email")
    idx_role = headers.index("role")
    
    # Optional columns map
    idx_dept = headers.index("departmentcode") if "departmentcode" in headers else -1
    idx_sect = headers.index("sectioncode") if "sectioncode" in headers else -1
    idx_desg = headers.index("designationcode") if "designationcode" in headers else -1

    imported_count = 0
    errors = []

    for r_num, row in enumerate(rows[1:], start=2):
        if not row or all(v is None for v in row):
            continue  # Skip blank rows

        name_val = str(row[idx_name]).strip() if row[idx_name] is not None else ""
        email_val = str(row[idx_email]).strip() if row[idx_email] is not None else ""
        role_val = str(row[idx_role]).strip() if row[idx_role] is not None else ""

        # Validation checks
        if not name_val or not email_val or not role_val:
            errors.append(f"Row {r_num}: Missing required field values.")
            continue

        email_regex = r"^[\w\.-]+@[\w\.-]+\.\w+$"
        if not re.match(email_regex, email_val):
            errors.append(f"Row {r_num}: Malformed email address structure.")
            continue

        if role_val == "MASTER_ADMIN":
            errors.append(f"Row {r_num}: Attempting to duplicate MASTER_ADMIN.")
            continue

        role = db.query(Role).filter_by(name=role_val).first()
        if not role:
            errors.append(f"Row {r_num}: Role '{role_val}' is invalid.")
            continue

        if db.query(User).filter_by(email=email_val).first():
            errors.append(f"Row {r_num}: Email already registered.")
            continue

        # Resolve optional parameters
        dept_id = None
        if idx_dept != -1 and row[idx_dept] is not None:
            dept_code = str(row[idx_dept]).strip()
            dept = db.query(Department).filter_by(code=dept_code).first()
            if dept:
                dept_id = dept.id
            else:
                errors.append(f"Row {r_num}: Department code '{dept_code}' invalid.")
                continue

        sect_id = None
        if idx_sect != -1 and row[idx_sect] is not None:
            sect_code = str(row[idx_sect]).strip()
            sect = db.query(Section).filter_by(code=sect_code).first()
            if sect:
                sect_id = sect.id
            else:
                errors.append(f"Row {r_num}: Section code '{sect_code}' invalid.")
                continue

        desg_id = None
        if idx_desg != -1 and row[idx_desg] is not None:
            desg_code = str(row[idx_desg]).strip()
            desg = db.query(Designation).filter_by(code=desg_code).first()
            if not desg:
                errors.append(f"Row {r_num}: Designation code '{desg_code}' invalid.")
                continue
            else:
                desg_id = desg.id

        # Generate credentials
        username = generate_username(name_val)
        while db.query(User).filter_by(username=username).first():
            username = generate_username(name_val)

        temp_pwd = generate_temp_password()
        hashed = get_password_hash(temp_pwd)

        uid = str(uuid.uuid4())
        user = User(
            id=uid,
            email=email_val,
            username=username,
            passwordHash=hashed,
            name=name_val,
            roleId=role.id,
            departmentId=dept_id,
            sectionId=sect_id,
            status="ACTIVE",
            mustChangePassword=True
        )
        db.add(user)

        # Create Profile
        profile = UserProfile(id=str(uuid.uuid4()), userId=uid, designationId=desg_id)
        db.add(profile)
        
        imported_count += 1
        print(f"[IMPORT LOG] Imported: {username} ({email_val}) / Password: {temp_pwd}")

    db.commit()

    # Log action to audit trail
    audit = AuditLog(
        id=str(uuid.uuid4()),
        userId=current_user.id,
        action="USERS_IMPORT",
        details=f"Successfully imported {imported_count} accounts. Failures: {len(errors)}."
    )
    db.add(audit)
    db.commit()

    import_data = {
        "imported": imported_count,
        "errors": errors
    }
    return make_response(
        success=True,
        message=f"Import completed. Imported: {imported_count}.",
        data=import_data,
        extra_compat=import_data
    )

@router.get("/export/xlsx", dependencies=[Depends(PermissionChecker("users:read"))])
def export_users(db: Session = Depends(get_db)):
    """Generates and streams an Excel sheet compile of all registered system users."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Campus Users Directory"

    headers = [
        "ID", "Name", "Username", "Email Address", "System Role", 
        "Academic Department", "Section Unit", "Account Status", "Created Date"
    ]
    ws.append(headers)

    # Styles
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True)

    users = db.query(User).all()
    for u in users:
        row = [
            u.id,
            u.name,
            u.username,
            u.email,
            u.role.name,
            u.department.name if u.department else "None",
            u.section.name if u.section else "None",
            u.status,
            u.createdAt.strftime("%Y-%m-%d %H:%M:%S")
        ]
        ws.append(row)

    # Save spreadsheet to bytes stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_directory.xlsx"}
    )
